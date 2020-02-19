from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

"""Проверка строки"""
def check(stroka):
    if stroka[len(stroka)-1:len(stroka)] == '.0':
        stroka
    if stroka.isdigit():
        chislo = int(stroka)
        answer1 = 'int'
        return answer1, chislo

    else:
        answer1 = 'str'
        return answer1, stroka


'''
Считывание таблицы
'''
answer = 'n'
while answer == 'н' or answer == 'n' or answer == 'not' or answer == 'нет':
    file_obj = input('Введите путь к файлу: ')
    workbook = load_workbook(filename=file_obj)
    sheet = workbook.worksheets[0]
    '''
    Проверка правильности считывания
    '''
    print('Количество строк в таблице: ', str(sheet.max_row))
    print('Количество столбцов в таблице: ', str(sheet.max_column))

    answer = input('Верно?\n')
    if answer == 'yes' or answer == 'y' or answer == 'да' or answer == 'д':
        print('Чтение файла завершилось верно')
        break
    else:
        print('Ошибка чтения файла!!!')
        continue

max_row = sheet.max_row
max_column = sheet.max_column
column_2 = column_index_from_string(input("Введите литер столбца, с которого начинаются данные гаплотипов: ").upper())

'''
Получение дополнительных данных
'''
booling = 'н'
while booling == 'н' or booling == 'n' or booling == 'not' or booling == 'нет':
    base_index = input('Введите индекс гаплотипа: ')
    type1, base_index_1 = check(base_index)
    for i in range(2, sheet.max_row + 1):
        type2, base_index_2 = check(str(sheet.cell(row=i, column=1).value))
        if type1 == type2 and base_index_1 == base_index_2:
            base_index = i
    print('Проверьте данные:')
    print('|  ', end='')
    print(base_index_1, end='   | ')
    for column1 in range(column_2, sheet.max_column + 1):
        if sheet.cell(row=base_index, column=column1).value is not None:
            print('{: ^ 4} | '.format(sheet.cell(row=base_index, column=column1).value), end='')
        else:
            continue
    print()
    booling = input("Правильно ?\n ")
    if booling == 'yes' or booling == 'y' or booling == 'да' or booling == 'д':
        break
age = float(input('Введите средний возраст покаления: '))
'''
const
'''
mu = 0.0026
'''
Подсчитаем TMRCA в гаплоидных наборах относительно базового 
'''
new_column = sheet.max_column + 1
title_coordinate = get_column_letter(new_column) + str(1)

sheet[title_coordinate] = 0
sheet[get_column_letter(sheet.max_column)+str(base_index)] = 0
for row1 in range(2, sheet.max_row + 1):
    diff_square = 0
    count_locus = 0
    if row1 != base_index:
        for column1 in range(column_2, sheet.max_column + 1):
            a = sheet.cell(row=base_index, column=column1).value
            b = sheet.cell(row=row1, column=column1).value
            if a is not None and b is not None:
                calculation = ((a - int(b)) ** 2) / 2
                diff_square += calculation
                count_locus += 1
            else:
                continue
    else:
        continue
    if count_locus != 0:
        average_value = diff_square / count_locus
        TMRCA = age * (average_value / mu)
        coordinate = get_column_letter(new_column) + str(row1)
        sheet[coordinate] = TMRCA
    else:
        continue
workbook.save(file_obj)

'''
Создаём список с данными
'''
sheet_of_project = []
for row1 in range(1, sheet.max_row+1):
    stroka = []
    for column1 in range(1, sheet.max_column + 1):
        if sheet.cell(row=row1, column=column1).value is not None:
            if column1 == sheet.max_column:
                stroka.append(float(sheet.cell(row=row1, column=column1).value))
            else:
                stroka.append(sheet.cell(row=row1, column=column1).value)
    if len(stroka) == sheet.max_column:
        sheet_of_project.append(stroka)
    else:
        continue
'''
Столбец относительно которого сортируем таблицу
'''
base_column = max_column
'''
Сортировка
'''

sheet_of_project.sort(key=lambda i: i[base_column])
'''
Вывод отсортированных данных в таблицу
'''
for row1 in range(0, len(sheet_of_project)):
    for column1 in range(0, sheet.max_column):
        coordinate = get_column_letter(column1+1) + str(row1+1)
        sheet[coordinate] = sheet_of_project[row1][column1]
sheet[title_coordinate] = 'TMRCA'
sheet[get_column_letter(sheet.max_column)+str(2)] = ''
workbook.save(file_obj)
